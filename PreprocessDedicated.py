import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

OFFSET = 0
file = ""
input_file = file+".csv"
output_file = file+"-off"+str(OFFSET)+".xlsx"
# Read CSV as plain text (one column first)
with open(input_file, "r", encoding="utf-8") as f:
    lines = f.readlines()
df = pd.DataFrame({"raw": [line.strip() for line in lines if line.strip()]})

raw_series = df.fillna("").astype(str).agg(" ".join, axis=1)
print(raw_series.shape)
df = pd.DataFrame({"raw": raw_series})

# Split into parts
df["Time Stamp"] = df["raw"].str.extract(r"\[(.*?)\]")
df["Time Stamp"] = pd.to_numeric(df["Time Stamp"], errors="coerce")
df["Title"] = df["raw"].str.extract(r"\](.*?):")
df["Description"] = df["raw"].str.split(":", n=1).str[1]
df["Aligned"] = df["Time Stamp"] + OFFSET

def fmt_mmss(t):
    if pd.isna(t):
        return ""
    sign = "-" if t < 0 else ""
    t = abs(float(t))
    m = int(t // 60)
    s = t - m * 60
    # Keep millisecond precision if fractional seconds exist
    out = f"{m:02d}:{s:06.3f}" if (s % 1) else f"{m:02d}:{int(s):02d}"
    # Trim trailing zeros if we used 3 decimals
    if "." in out:
        out = out.rstrip("0").rstrip(".")
    return sign + out

df["Aligned Time (mm:ss)"] = df["Aligned"].apply(fmt_mmss)
df["Time (mm:ss)"] = df["Time Stamp"].apply(fmt_mmss)

# Drop the raw column if no longer needed
df = df[["Time Stamp", "Aligned", "Aligned Time (mm:ss)", "Title", "Description"]]
print(df)
df_speech = df[df["Title"].str.strip().eq("SPEECH")]
df_eyegaze = df[df["Title"].str.strip().eq("EyeGaze")]
df_others = df[~df["Title"].str.strip().isin(["SPEECH", "EyeGaze"])]

#Parse EyeGaze
desc = df_eyegaze["Description"].fillna("").astype(str).str.strip()
parsed = desc.str.extract(r'^\s*([A-Za-z]+)\s*[_-]\s*([A-Za-z]+)\s*[_-]\s*([A-Za-z0-9]+)\b\s*(.*)$')
parsed.columns = ["left/right", "enter/exit", "Place", "extra description"]
parsed["left/right"] = parsed["left/right"].str.lower().replace({"l":"left","r":"right"})
parsed["enter/exit"] = parsed["enter/exit"].str.lower().replace({"in":"enter","out":"exit"})
df_eyegaze = pd.concat([df_eyegaze, parsed], axis=1)

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="All")
    df_speech.to_excel(writer, index=False, sheet_name="SPEECH")
    df_eyegaze.to_excel(writer, index=False, sheet_name="EyeGaze")
    df_others.to_excel(writer, index=False, sheet_name="Others")

wb = load_workbook(output_file)
for ws in wb.worksheets:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value = str(cell.value)
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        ws.column_dimensions[col_letter].width = (max_length + 2)

    # Apply AutoFilter on header row
    last_col = get_column_letter(ws.max_column)
    last_row = ws.max_row
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Freeze top row
    ws.freeze_panes = "A2"

wb.save(output_file)
